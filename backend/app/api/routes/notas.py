from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, extract
from typing import Optional
from datetime import datetime
import io
import zipfile
import os

from app.db.database import get_db
from app.models.models import Nota, Empresa, NotaTipo, NotaModelo, NotaStatus
from app.schemas.schemas import NotaResponse, NotaListResponse, EstatisticasResponse
from app.core.deps import get_current_user, get_current_empresa, require_active_empresa
from app.models.models import Usuario

router = APIRouter()


@router.get("", response_model=NotaListResponse)
async def listar_notas(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    tipo: Optional[str] = None,
    modelo: Optional[str] = None,
    status: Optional[str] = None,
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    query = select(Nota).where(Nota.empresa_id == empresa.id)

    if data_inicio:
        query = query.where(Nota.data_emissao >= data_inicio)
    if data_fim:
        query = query.where(Nota.data_emissao <= data_fim)
    if tipo:
        query = query.where(Nota.tipo == tipo)
    if modelo:
        query = query.where(Nota.modelo == modelo)
    if status:
        query = query.where(Nota.status == status)

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar()

    # Paginate
    query = query.order_by(Nota.criado_em.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = result.scalars().all()

    return NotaListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size if total > 0 else 0
    )


@router.get("/estatisticas")
async def estatisticas(
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    now = datetime.utcnow()
    mes_atual = now.month
    ano_atual = now.year

    base = and_(
        Nota.empresa_id == empresa.id,
        extract('month', Nota.data_emissao) == mes_atual,
        extract('year', Nota.data_emissao) == ano_atual
    )

    # Entrada count
    r = await db.execute(select(func.count(Nota.id)).where(and_(base, Nota.tipo == "entrada")))
    total_entrada = r.scalar() or 0

    # Saida count
    r = await db.execute(select(func.count(Nota.id)).where(and_(base, Nota.tipo == "saida")))
    total_saida = r.scalar() or 0

    # Canceladas
    r = await db.execute(select(func.count(Nota.id)).where(and_(
        Nota.empresa_id == empresa.id,
        Nota.status == "cancelada"
    )))
    total_canceladas = r.scalar() or 0

    # Valor total mes
    r = await db.execute(select(func.sum(Nota.valor_total)).where(base))
    valor_total = r.scalar() or 0

    # Monthly chart - last 6 months
    grafico = []
    for i in range(5, -1, -1):
        m = (mes_atual - i - 1) % 12 + 1
        a = ano_atual if mes_atual - i > 0 else ano_atual - 1
        r = await db.execute(select(func.count(Nota.id)).where(and_(
            Nota.empresa_id == empresa.id,
            extract('month', Nota.data_emissao) == m,
            extract('year', Nota.data_emissao) == a
        )))
        cnt = r.scalar() or 0
        grafico.append({"mes": f"{m:02d}/{a}", "total": cnt})

    return {
        "total_entrada_mes": total_entrada,
        "total_saida_mes": total_saida,
        "total_canceladas": total_canceladas,
        "valor_total_mensal": float(valor_total),
        "grafico_mensal": grafico
    }


@router.get("/exportar")
async def exportar_excel(
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    result = await db.execute(
        select(Nota).where(Nota.empresa_id == empresa.id).order_by(Nota.data_emissao.desc()).limit(10000)
    )
    notas = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    headers = ["Chave", "Modelo", "Tipo", "Emitente", "Destinatário", "Valor", "Emissão", "Status"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")

    for row, nota in enumerate(notas, 2):
        ws.cell(row=row, column=1, value=nota.chave)
        ws.cell(row=row, column=2, value=nota.modelo)
        ws.cell(row=row, column=3, value=nota.tipo)
        ws.cell(row=row, column=4, value=nota.cnpj_emitente)
        ws.cell(row=row, column=5, value=nota.cnpj_destinatario)
        ws.cell(row=row, column=6, value=float(nota.valor_total) if nota.valor_total else 0)
        ws.cell(row=row, column=7, value=nota.data_emissao.strftime("%d/%m/%Y") if nota.data_emissao else "")
        ws.cell(row=row, column=8, value=nota.status)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=notas_fiscais.xlsx"}
    )


@router.get("/download-lote")
async def download_lote(
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Nota).where(
            and_(Nota.empresa_id == empresa.id, Nota.xml_path != None)
        ).limit(500)
    )
    notas = result.scalars().all()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for nota in notas:
            if nota.xml_path and os.path.exists(nota.xml_path):
                zip_file.write(nota.xml_path, f"{nota.chave}.xml")

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=xmls_fiscais.zip"}
    )


@router.get("/download/{nota_id}")
async def download_xml(
    nota_id: str,
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Nota).where(and_(Nota.id == nota_id, Nota.empresa_id == empresa.id))
    )
    nota = result.scalar_one_or_none()
    if not nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    if not nota.xml_path or not os.path.exists(nota.xml_path):
        raise HTTPException(status_code=404, detail="XML não disponível")

    return FileResponse(nota.xml_path, media_type="application/xml", filename=f"{nota.chave}.xml")


@router.get("/{nota_id}", response_model=NotaResponse)
async def get_nota(
    nota_id: str,
    empresa: Empresa = Depends(require_active_empresa),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Nota).where(and_(Nota.id == nota_id, Nota.empresa_id == empresa.id))
    )
    nota = result.scalar_one_or_none()
    if not nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    return nota
